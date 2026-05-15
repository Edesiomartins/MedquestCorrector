"""Exporta resultados de correção para planilha Excel."""

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def humanize_review_reason(
    warnings: list[str] | None,
    technical_error: str | None = None,
    flags: dict[str, Any] | None = None,
) -> str:
    flags = flags or {}
    combined = " | ".join([*(warnings or []), technical_error or ""]).lower()
    if flags.get("empty_answer") or "sem resposta" in combined or "resposta vazia" in combined:
        return "Resposta em branco ou não detectada."
    if "cópia do enunciado" in combined or "copia do enunciado" in combined:
        return "Resposta parece cópia do enunciado da questão."
    if "schema incompleto" in combined or "json" in combined:
        return "Falha no formato da resposta da IA. Conferir nota sugerida."
    if "expecting property name" in combined:
        return "Falha ao interpretar resposta da IA. Revisão manual recomendada."
    if "identity_source" in combined or "manifest_fallback" in combined or "sem qr confiável" in combined:
        return "Vinculação do aluno feita sem QR confiável. Conferir aluno."
    if flags.get("low_confidence") or "baixa confiança" in combined:
        return "Baixa confiança na leitura da resposta manuscrita."
    if "off-topic" in combined or "fora do tema" in combined:
        return "Resposta fora do tema da questão."
    if combined.strip():
        return "Revisão manual recomendada."
    return ""


def _round_quarter(value: float) -> float:
    return round(value * 4) / 4


def _sanitize_sheet_title_fragment(text: str) -> str:
    cleaned = str(text or "")
    for ch in r"[]*?:/\\":
        cleaned = cleaned.replace(ch, " ")
    cleaned = " ".join(cleaned.split()).strip()
    return cleaned or "Aluno"


def _unique_sheet_title(registration: str, student_name: str, used_titles: set[str]) -> str:
    base = _sanitize_sheet_title_fragment(f"{registration} {student_name}")[:31]
    if not base:
        base = "Aluno"
    candidate = base[:31]
    suffix_n = 2
    while candidate in used_titles:
        suffix = f" ({suffix_n})"
        candidate = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else (base + suffix)
        candidate = candidate[:31]
        suffix_n += 1
    used_titles.add(candidate)
    return candidate


def _final_score_for_row(
    r: dict,
    questions: list[dict],
    *,
    round_final_to_quarter: bool,
) -> float:
    scores = r.get("scores", {})
    numeric_scores = [float(v) for v in scores.values() if isinstance(v, (int, float))]
    final_score = sum(numeric_scores) / len(questions) if questions else 0.0
    if round_final_to_quarter:
        final_score = _round_quarter(final_score)
    return final_score


def export_results_xlsx(
    exam_name: str,
    questions: list[dict],
    results: list[dict],
    *,
    include_details: bool = True,
    round_final_to_quarter: bool = True,
) -> bytes:
    """
    Gera um arquivo .xlsx com as notas.

    - Aba ``Resultado Final``: uma linha por aluno (resumo).
    - Com ``include_details=True``: uma aba por aluno (dados completos + tabela de questões)
      e aba ``Revisões Necessárias``.

    questions: [{"number": 1, "text": "...", "max_score": 1.0, "expected_answer": "..."}, ...]
    results: [{
        "student_name": str,
        "registration_number": str,
        "curso": str,
        "turma": str,
        "scores": {1: 0.75, 2: 1.0, ...},  # question_number -> final_score
        "total": float,
    }, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado Final"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    question_lookup = {
        int(q.get("number")): q
        for q in questions
        if isinstance(q.get("number"), int) or str(q.get("number") or "").isdigit()
    }

    headers = ["Matrícula", "Nome", "Curso", "Turma", "Nota final", "Revisão necessária", "Motivo da revisão"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["registration_number"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=r["student_name"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=r.get("curso", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value=r.get("turma", "")).border = thin_border

        final_score = _final_score_for_row(r, questions, round_final_to_quarter=round_final_to_quarter)
        review_reason = humanize_review_reason(
            r.get("warnings") or [],
            r.get("technical_error") or r.get("observacoes") or "",
            {"low_confidence": r.get("low_confidence"), "empty_answer": r.get("empty_answer")},
        )

        total_col = 5
        total_cell = ws.cell(row=row_idx, column=total_col, value=final_score)
        total_cell.font = Font(bold=True)
        total_cell.alignment = center
        total_cell.border = thin_border

        review_cell = ws.cell(
            row=row_idx,
            column=total_col + 1,
            value="sim" if r.get("needs_review") else "não",
        )
        review_cell.alignment = center
        review_cell.border = thin_border

        obs_cell = ws.cell(row=row_idx, column=total_col + 2, value=review_reason)
        obs_cell.border = thin_border

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["G"].width = 44

    # Título acima
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=exam_name)
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    if include_details:
        used_titles = set(wb.sheetnames)
        used_titles.add("Revisões Necessárias")
        for r in results:
            _add_student_sheet(
                wb,
                exam_name,
                r,
                question_lookup,
                questions,
                thin_border,
                header_font,
                header_fill,
                center,
                used_titles,
                round_final_to_quarter=round_final_to_quarter,
            )
        _add_review_sheet(wb, results, question_lookup, thin_border, header_font, header_fill, center)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_header(ws, headers, header_font, header_fill, thin_border, center):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def _expected_answer_for(detail: dict, question_lookup: dict[int, dict]) -> str:
    direct = detail.get("expected_answer") or detail.get("answer_key") or detail.get("rubric_expected_answer")
    if direct:
        return str(direct)
    try:
        question_number = int(detail.get("question_number"))
    except (TypeError, ValueError):
        return ""
    question = question_lookup.get(question_number) or {}
    return str(question.get("expected_answer") or question.get("answer_key") or "")


def _add_student_sheet(
    wb: Workbook,
    exam_name: str,
    r: dict,
    question_lookup: dict[int, dict],
    questions: list[dict],
    thin_border: Border,
    header_font: Font,
    header_fill: PatternFill,
    center: Alignment,
    used_titles: set[str],
    *,
    round_final_to_quarter: bool,
) -> None:
    """Uma planilha por aluno: identificação + tabela completa das questões."""
    title = _unique_sheet_title(
        str(r.get("registration_number", "")),
        str(r.get("student_name", "")),
        used_titles,
    )
    ws = wb.create_sheet(title)

    thin = thin_border
    section_font = Font(bold=True, size=11)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    t1 = ws.cell(row=1, column=1, value=str(exam_name or "Prova"))
    t1.font = Font(bold=True, size=14)
    t1.alignment = Alignment(horizontal="left", vertical="center")

    ws.cell(row=2, column=1, value=str(r.get("student_name") or "")).font = Font(bold=True, size=12)

    label_font = Font(bold=True, size=10)
    row = 4
    ws.cell(row=row, column=1, value="Dados gerais").font = section_font
    row += 1

    aggregated_review = humanize_review_reason(
        r.get("warnings") or [],
        r.get("technical_error") or r.get("observacoes") or "",
        {"low_confidence": r.get("low_confidence"), "empty_answer": r.get("empty_answer")},
    )

    info_rows = [
        ("Matrícula", r.get("registration_number", "")),
        ("Nome", r.get("student_name", "")),
        ("Curso", r.get("curso", "")),
        ("Turma", r.get("turma", "")),
        ("Nota final", _final_score_for_row(r, questions, round_final_to_quarter=round_final_to_quarter)),
        ("Total (somatório pontos)", r.get("total", "")),
        ("Revisão necessária (prova inteira)", "sim" if r.get("needs_review") else "não"),
        ("Motivo revisão (agregado)", aggregated_review),
        ("Origem identidade", str(r.get("identity_source") or "")),
        ("Observações", str(r.get("observacoes") or "")),
    ]
    for label, val in info_rows:
        la = ws.cell(row=row, column=1, value=label)
        la.font = label_font
        la.border = thin
        vc = ws.cell(row=row, column=2, value=val)
        vc.border = thin
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Detalhamento por questão").font = section_font
    row += 1

    q_headers = [
        "Questão",
        "Nota",
        "Veredito",
        "Comentário",
        "Resposta esperada",
        "Transcrição",
        "Revisão necessária",
        "Motivo da revisão",
        "Página física",
        "Confiança da transcrição",
    ]
    header_row = row
    for col_idx, h in enumerate(q_headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    details = list(r.get("question_details") or [])

    def _qnum(d: dict) -> int:
        try:
            return int(d.get("question_number"))
        except (TypeError, ValueError):
            return 0

    details.sort(key=_qnum)
    data_start = header_row + 1
    for i, detail in enumerate(details):
        rn = data_start + i
        tc = detail.get("transcription_confidence")
        low_conf = False
        if tc is not None:
            try:
                low_conf = float(tc) < 0.70
            except (TypeError, ValueError):
                low_conf = False
        friendly = humanize_review_reason(
            detail.get("warnings") or [],
            detail.get("technical_detail") or detail.get("review_reason") or "",
            {
                "low_confidence": low_conf,
                "empty_answer": not str(detail.get("transcription") or "").strip(),
            },
        )
        vals = [
            detail.get("question_number"),
            detail.get("score"),
            detail.get("verdict", ""),
            detail.get("comment", ""),
            _expected_answer_for(detail, question_lookup),
            detail.get("transcription", ""),
            "sim" if detail.get("needs_review") else "não",
            friendly,
            detail.get("physical_page"),
            detail.get("transcription_confidence"),
        ]
        for col_idx, value in enumerate(vals, 1):
            ws.cell(row=rn, column=col_idx, value=value).border = thin

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["D"].width = 44
    ws.column_dimensions["E"].width = 48
    ws.column_dimensions["F"].width = 48
    ws.column_dimensions["H"].width = 40


def _add_review_sheet(wb, results, question_lookup, thin_border, header_font, header_fill, center):
    ws = wb.create_sheet("Revisões Necessárias")
    headers = [
        "Matrícula",
        "Nome",
        "Questão",
        "Nota sugerida",
        "Resposta esperada",
        "Motivo amigável",
        "Detalhe técnico resumido",
        "Página física",
        "Confiança da transcrição",
        "Identity source",
    ]
    _write_header(ws, headers, header_font, header_fill, thin_border, center)
    row_idx = 2
    for r in results:
        for detail in r.get("question_details", []):
            if not detail.get("needs_review"):
                continue
            friendly = humanize_review_reason(
                detail.get("warnings") or [],
                detail.get("technical_detail") or detail.get("review_reason") or "",
                {
                    "low_confidence": (detail.get("transcription_confidence") is not None and detail.get("transcription_confidence") < 0.70),
                    "empty_answer": not (detail.get("transcription") or "").strip(),
                },
            ) or "Revisão manual recomendada."
            values = [
                r.get("registration_number", ""),
                r.get("student_name", ""),
                detail.get("question_number"),
                detail.get("score"),
                _expected_answer_for(detail, question_lookup),
                friendly,
                str(detail.get("technical_detail") or detail.get("review_reason") or "")[:500],
                detail.get("physical_page"),
                detail.get("transcription_confidence"),
                r.get("identity_source", ""),
            ]
            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=value).border = thin_border
            row_idx += 1
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 44
    ws.column_dimensions["G"].width = 60
